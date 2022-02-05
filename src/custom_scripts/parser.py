import os
from item_master.models import Item
# Import the openpyxl module
import yaml

# since it is running from django shell cwd is not working
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = '/home/aj355y/workspace/opensource/rm/croma/src/custom_scripts'
file_name = os.path.join(BASE_DIR, 'drug.xlsx')

def import_product():
    from openpyxl import load_workbook
    wb = load_workbook(filename = file_name)
    ws = wb['List of Registered Medicines']
    names = list()
    data = []
    i=3
    for cell in ws['i']:
        if cell.value and cell.value != 'n/a':
            value = cell.value.strip('\n').strip().encode('ascii', 'ignore').decode("utf-8").strip()
            if value.lower() not in names:
                # login to skip duplicate
                names.append(value.lower())
                Item.objects.create(name=value)
                i +=1


#import_manufacturer()
import_product()